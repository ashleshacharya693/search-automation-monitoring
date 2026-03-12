import random
import os
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ==============================
# 🔹 SET RANDOM PLATFORM & USER BEFORE COLLECTION
# ==============================

# ==============================
# 🔹 CUSTOM CLI OPTIONS
# ==============================
def pytest_addoption(parser):
    parser.addoption("--platform", action="store", default=None, help="Platform to test (web/app/tv)")
    parser.addoption("--user", action="store", default=None, help="User type to test")

def pytest_configure(config):
    from config.platforms import PLATFORMS
    from config.users import USERS

    platform = random.choice(list(PLATFORMS.keys()))
    user     = random.choice(list(USERS.keys()))

    os.environ["FUZZY_PLATFORM"] = platform
    os.environ["FUZZY_USER"]     = user

    print(f"\n[conftest] Fuzzy test → Platform: {platform} | User: {user}")



# ==============================
# 🔹 SHARED HELPERS
# ==============================

def _write_sheet(ws, results, title):
    """Write a results list to a given worksheet with formatting + summary."""

    if not results:
        ws.cell(row=1, column=1, value=f"No {title} results found.")
        return None

    headers = list(results[0].keys())

    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="343A40", fill_type="solid")

    # Data rows
    for row_num, row_data in enumerate(results, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data[header])
            if row_data["Status"] == "FAILED":
                cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Summary block
    total    = len(results)
    passed   = len([r for r in results if r["Status"] == "PASSED"])
    failed   = total - passed
    accuracy = (passed / total) * 100 if total > 0 else 0

    summary_start = total + 3

    cell = ws.cell(row=summary_start, column=1, value="Total Tests:")
    cell.font = Font(bold=True)
    ws.cell(row=summary_start, column=2, value=total)

    cell = ws.cell(row=summary_start + 1, column=1, value="Passed:")
    cell.font = Font(bold=True)
    ws.cell(row=summary_start + 1, column=2, value=passed)

    cell = ws.cell(row=summary_start + 2, column=1, value="Failed:")
    cell.font = Font(bold=True)
    ws.cell(row=summary_start + 2, column=2, value=failed)

    cell = ws.cell(row=summary_start + 3, column=1, value="Accuracy %:")
    cell.font = Font(bold=True)
    ws.cell(row=summary_start + 3, column=2, value=round(accuracy, 2))

    return total, passed, failed, accuracy


def _save_report(results, sheet_title, filename, label):
    """Create a workbook, write results, save and print summary."""

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    stats = _write_sheet(ws, results, label)
    wb.save(filename)

    if stats:
        total, passed, failed, accuracy = stats
        print(f"\n{'='*55}")
        print(f"  {label} Report → {filename}")
        print(f"  Total: {total} | Passed: {passed} | Failed: {failed}")
        print(f"  Accuracy: {accuracy:.2f}%")
        print(f"{'='*55}\n")


# ==============================
# 🔹 SESSION FINISH HOOK
# ==============================

def pytest_sessionfinish(session, exitstatus):

    # ── Premium report ────────────────────────────────────────────────────────
    premium_results = getattr(pytest, "results_summary", [])
    if premium_results:
        _save_report(
            results     = premium_results,
            sheet_title = "Premium Report",
            filename    = "reports/premium_report.xlsx",
            label       = "Premium",
        )
    else:
        print("\n⚠️  No premium results to report.")

    # ── Live Match report ─────────────────────────────────────────────────────
    live_results = getattr(pytest, "live_results_summary", [])
    if live_results:
        _save_report(
            results     = live_results,
            sheet_title = "Live Match Report",
            filename    = "reports/live_match_report.xlsx",
            label       = "Live Match",
        )
    else:
        print("\n⚠️  No live match results to report.")

    # ── Offboarded report ─────────────────────────────────────────────────────
    offboarded_results = getattr(pytest, "offboarded_results_summary", [])
    if offboarded_results:
        _save_report(
            results     = offboarded_results,
            sheet_title = "Offboarded Report",
            filename    = "reports/offboarded_report.xlsx",
            label       = "Offboarded",
        )
    else:
        print("\n⚠️  No offboarded results to report.")

    # ── Fuzzy & Synonym report ────────────────────────────────────────────────
    fuzzy_results = getattr(pytest, "fuzzy_results_summary", [])
    if fuzzy_results:
        _save_report(
            results     = fuzzy_results,
            sheet_title = "Fuzzy & Synonym Report",
            filename    = "reports/fuzzy_synonym_report.xlsx",
            label       = "Fuzzy & Synonym",
        )
    else:
        print("\n⚠️  No fuzzy/synonym results to report.")

    # ── Sport Tournament report ───────────────────────────────────────────────
    tournament_results = getattr(pytest, "tournament_results_summary", [])
    if tournament_results:
        _save_report(
            results     = tournament_results,
            sheet_title = "Tournament Match Coverage",
            filename    = "reports/sport_tournament_report.xlsx",
            label       = "Sport Tournament",
        )
    else:
        print("\n⚠️  No tournament match results to report.")

    # ── Empty Tournament report ───────────────────────────────────────────────
    empty_tournament_results = getattr(pytest, "empty_tournament_results_summary", [])
    if empty_tournament_results:
        _save_report(
            results     = empty_tournament_results,
            sheet_title = "Empty Tournament Check",
            filename    = "reports/empty_tournament_report.xlsx",
            label       = "Empty Tournament",
        )
    else:
        print("\n⚠️  No empty tournament results to report.")